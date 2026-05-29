"""
NSE Nifty LargeMidCap 250 Daily DMA Screener
=============================================
Fixed: Yahoo Finance ticker mapping for all NSE symbols
"""

import os, io, json, time, zipfile, warnings
import pandas as pd
import requests
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Config ─────────────────────────────────────────────────────────────────────
DOCS_DIR                = "./docs"
HISTORY_DIR             = "./nse_history"
SYMBOLS_CACHE_FILE      = "./nse_history/nifty250_symbols_cache.json"
HIGH_VOL_MULTIPLIER     = 1.5
HIGH_VALUE_CRORE        = 5
DMA200_OVEREXTENDED_PCT = 10
HISTORY_DAYS            = 220

NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/",
    "Connection": "keep-alive",
}

# ══════════════════════════════════════════════════════════════════════
# YAHOO FINANCE TICKER MAP
# NSE Symbol → Yahoo Finance Ticker
# All known differences between NSE symbol and Yahoo .NS ticker
# ══════════════════════════════════════════════════════════════════════
YF_TICKER_MAP = {
    # Special characters
    "M&M":            "M%26M.NS",
    "BAJAJ-AUTO":     "BAJAJ-AUTO.NS",
    "MCDOWELL-N":     "MCDOWELL-N.NS",
    "L&TFH":          "L%26TFH.NS",

    # Name changes / different Yahoo tickers
    "LTIM":           "LTIM.NS",          # LTIMindtree
    "TATAMOTORS":     "TATAMOTORS.NS",    # standard
    "LODHA":          "LODHA.NS",         # Macrotech Developers
    "DMART":          "DMART.NS",         # Avenue Supermarts
    "LICI":           "LICI.NS",          # LIC India
    "BAJAJHFL":       "BAJAJHFL.NS",      # Bajaj Housing Finance
    "ATGL":           "ATGL.NS",          # Adani Total Gas
    "GMRAIRPORT":     "GMRAIRPORT.NS",    # GMR Airports
    "PATANJALI":      "PATANJALI.NS",     # Patanjali Foods
    "PRESTIGE":       "PRESTIGE.NS",      # Prestige Estates
    "PVRINOX":        "PVRINOX.NS",       # PVR Inox
    "NSLNISP":        "NSLNISP.NS",
    "RKFORGE":        "RKFORGE.NS",
    "ROUTE":          "ROUTE.NS",
    "INDIANB":        "INDIANB.NS",
    "ANANDRATHI":     "ANANDRATHI.NS",
    "BSE":            "BSE.NS",
    "IEX":            "IEX.NS",
    "METROBRAND":     "METROBRAND.NS",    # Metro Brands
    "VOLTAMP":        "VOLTAMP.NS",
    "VODAFONE":       "IDEA.NS",          # Vodafone Idea trades as IDEA
    "APLAPOLLO":      "APLAPOLLO.NS",
    "TRITURBINE":     "TRITURBINE.NS",
    "STLTECH":        "STLTECH.NS",
    "TANLA":          "TANLA.NS",
    "OLECTRA":        "OLECTRA.NS",
    "SAFARI":         "SAFARI.NS",
    "GPIL":           "GPIL.NS",
    "JSL":            "JSL.NS",
    "JSWENERGY":      "JSWENERGY.NS",
    "CGPOWER":        "CGPOWER.NS",
    "SUMICHEM":       "SUMICHEM.NS",
    "KRISHANA":       "KRISHANA.NS",
    "PGHH":           "PGHH.NS",
    "GODFRYPHLP":     "GODFRYPHLP.NS",
    "NIACL":          "NIACL.NS",
    "NLCINDIA":       "NLCINDIA.NS",
    "TITAGARH":       "TITAGARH.NS",
    "IIFL":           "IIFL.NS",
    "KPITTECH":       "KPITTECH.NS",
    "TEAMLEASE":      "TEAMLEASE.NS",
    "MOTILALOFS":     "MOTILALOFS.NS",
    "SPARC":          "SPARC.NS",
    "POONAWALLA":     "POONAWALLA.NS",
    "UJJIVANSFB":     "UJJIVANSFB.NS",
    "UCOBANK":        "UCOBANK.NS",
    "VMART":          "VMART.NS",
    "TCNSBRANDS":     "TCNSBRANDS.NS",
    "TATAINVEST":     "TATAINVEST.NS",
    "MTAR":           "MTAR.NS",
    "ISEC":           "ISEC.NS",
    "FLUOROCHEM":     "FLUOROCHEM.NS",
    "ELGIEQUIP":      "ELGIEQUIP.NS",
    "DELHIVERY":      "DELHIVERY.NS",
    "CHAMBLFERT":     "CHAMBLFERT.NS",
    "BLUESTARCO":     "BLUESTARCO.NS",
    "ENDURANCE":      "ENDURANCE.NS",
    "EXIDEIND":       "EXIDEIND.NS",
    "FIVESTAR":       "FIVESTAR.NS",
    "BRIGADE":        "BRIGADE.NS",
    "SAREGAMA":       "SAREGAMA.NS",
    "KRBL":           "KRBL.NS",
    "SOBHA":          "SOBHA.NS",
    "REDINGTON":      "REDINGTON.NS",
    "RAILTEL":        "RAILTEL.NS",
    "UNOMINDA":       "UNOMINDA.NS",
    "MAPMYINDIA":     "MAPMYINDIA.NS",
    "MEDANTA":        "MEDANTA.NS",
    "MASTEK":         "MASTEK.NS",
    "KTKBANK":        "KTKBANK.NS",
    "KARURVYSYA":     "KARURVYSYA.NS",
    "AJANTPHARMA":    "AJANTPHARMA.NS",
    "AARTI":          "AARTIIND.NS",      # Aarti Industries trades as AARTIIND
    "BHEL":           "BHEL.NS",
    "AIAENG":         "AIAENG.NS",
    "AAVAS":          "AAVAS.NS",
    "APTUS":          "APTUS.NS",
    "ASTRAL":         "ASTRAL.NS",
    "ATUL":           "ATUL.NS",
    "AUROPHARMA":     "AUROPHARMA.NS",
    "BIOCON":         "BIOCON.NS",
    "CAMS":           "CAMS.NS",
    "CANFINHOME":     "CANFINHOME.NS",
    "CDSL":           "CDSL.NS",
    "CESC":           "CESC.NS",
    "COFORGE":        "COFORGE.NS",
    "CROMPTON":       "CROMPTON.NS",
    "DALBHARAT":      "DALBHARAT.NS",
    "DEEPAKNTR":      "DEEPAKNTR.NS",
    "DIXON":          "DIXON.NS",
    "EMAMILTD":       "EMAMILTD.NS",
    "ESCORTS":        "ESCORTS.NS",
    "FORTIS":         "FORTIS.NS",
    "GLAND":          "GLAND.NS",
    "GLAXO":          "GLAXO.NS",
    "GODREJIND":      "GODREJIND.NS",
    "GRINDWELL":      "GRINDWELL.NS",
    "HUDCO":          "HUDCO.NS",
    "IGL":            "IGL.NS",
    "INDHOTEL":       "INDHOTEL.NS",
    "INDIGO":         "INDIGO.NS",
    "IPCALAB":        "IPCALAB.NS",
    "JKCEMENT":       "JKCEMENT.NS",
    "JUBLFOOD":       "JUBLFOOD.NS",
    "KAJARIACER":     "KAJARIACER.NS",
    "KANSAINER":      "KANSAINER.NS",
    "KEC":            "KEC.NS",
    "KEI":            "KEI.NS",
    "LALPATHLAB":     "LALPATHLAB.NS",
    "LATENTVIEW":     "LATENTVIEW.NS",
    "LICHSGFIN":      "LICHSGFIN.NS",
    "LINDEINDIA":     "LINDEINDIA.NS",
    "MANAPPURAM":     "MANAPPURAM.NS",
    "MCX":            "MCX.NS",
    "MFSL":           "MFSL.NS",
    "MGL":            "MGL.NS",
    "MUTHOOTFIN":     "MUTHOOTFIN.NS",
    "NATCOPHARM":     "NATCOPHARM.NS",
    "NBCC":           "NBCC.NS",
    "NCC":            "NCC.NS",
    "NHPC":           "NHPC.NS",
    "NMDC":           "NMDC.NS",
    "OBEROIRLTY":     "OBEROIRLTY.NS",
    "PAGEIND":        "PAGEIND.NS",
    "PERSISTENT":     "PERSISTENT.NS",
    "PHOENIXLTD":     "PHOENIXLTD.NS",
    "PNBHOUSING":     "PNBHOUSING.NS",
    "POLYCAB":        "POLYCAB.NS",
    "PVRINOX":        "PVRINOX.NS",
    "RADICO":         "RADICO.NS",
    "RAMCOCEM":       "RAMCOCEM.NS",
    "RAYMOND":        "RAYMOND.NS",
    "RELAXO":         "RELAXO.NS",
    "RITES":          "RITES.NS",
    "RRKABEL":        "RRKABEL.NS",
    "SBICARD":        "SBICARD.NS",
    "SCHAEFFLER":     "SCHAEFFLER.NS",
    "SJVN":           "SJVN.NS",
    "SKFINDIA":       "SKFINDIA.NS",
    "SRF":            "SRF.NS",
    "STARHEALTH":     "STARHEALTH.NS",
    "SUNDARMFIN":     "SUNDARMFIN.NS",
    "SUPREMEIND":     "SUPREMEIND.NS",
    "SUNTV":          "SUNTV.NS",
    "SYNGENE":        "SYNGENE.NS",
    "TATACOMM":       "TATACOMM.NS",
    "TATAELXSI":      "TATAELXSI.NS",
    "TIMKEN":         "TIMKEN.NS",
    "TORNTPOWER":     "TORNTPOWER.NS",
    "UTIAMC":         "UTIAMC.NS",
    "VGUARD":         "VGUARD.NS",
    "WHIRLPOOL":      "WHIRLPOOL.NS",
    "YESBANK":        "YESBANK.NS",
    "ZEEL":           "ZEEL.NS",
    "ZYDUSLIFE":      "ZYDUSLIFE.NS",
    "ANGELONE":       "ANGELONE.NS",
    "KAYNES":         "KAYNES.NS",
    "BALKRISIND":     "BALKRISIND.NS",
    "BANDHANBNK":     "BANDHANBNK.NS",
    "FEDERALBNK":     "FEDERALBNK.NS",
    "ABCAPITAL":      "ABCAPITAL.NS",
    "ABFRL":          "ABFRL.NS",
    "ACC":            "ACC.NS",
    "DLF":            "DLF.NS",
    "ICICIGI":        "ICICIGI.NS",
    "IRFC":           "IRFC.NS",
}

def to_yf(sym):
    """Convert NSE symbol to Yahoo Finance ticker."""
    if sym in YF_TICKER_MAP:
        return YF_TICKER_MAP[sym]
    # Default: just append .NS
    return f"{sym}.NS"

# ══════════════════════════════════════════════════════════════════════
# HARDCODED FALLBACK (Tier 3 — emergency only)
# ══════════════════════════════════════════════════════════════════════
NIFTY_50 = [
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK",
    "BAJAJ-AUTO","BAJFINANCE","BAJAJFINSV","BEL","BHARTIARTL",
    "BPCL","BRITANNIA","CIPLA","COALINDIA","DRREDDY",
    "EICHERMOT","GRASIM","HCLTECH","HDFCBANK","HDFCLIFE",
    "HEROMOTOCO","HINDALCO","HINDUNILVR","ICICIBANK","INDUSINDBK",
    "INFY","ITC","JSWSTEEL","KOTAKBANK","LT",
    "LTIM","M&M","MARUTI","NESTLEIND","NTPC",
    "ONGC","POWERGRID","RELIANCE","SBILIFE","SBIN",
    "SHRIRAMFIN","SUNPHARMA","TATACONSUM","TATAMOTORS","TATASTEEL",
    "TCS","TECHM","TITAN","TRENT","ULTRACEMCO","WIPRO",
]
NIFTY_NEXT_50 = [
    "ABB","ADANIGREEN","ADANIPOWER","AMBUJACEM","BAJAJHFL",
    "BANKBARODA","BSE","CANBK","CHOLAFIN","COLPAL",
    "DABUR","DMART","GAIL","GODREJCP","HAVELLS",
    "HDFCAMC","HINDPETRO","ICICIGI","ICICIPRULI","INDUSTOWER",
    "IOC","IRCTC","IRFC","LICI","LODHA",
    "MARICO","MCDOWELL-N","MOTHERSON","MPHASIS","NAUKRI",
    "NMDC","OFSS","OIL","PAGEIND","PFC",
    "PIDILITIND","PIIND","RECLTD","SAIL","SIEMENS",
    "SRF","TATAPOWER","TORNTPHARM","TVSMOTOR","UBL",
    "UNIONBANK","VBL","VEDL","ZOMATO","ZYDUSLIFE",
]
NIFTY_MIDCAP_150 = [
    "AARTI","AAVAS","ABCAPITAL","ABFRL","ACC",
    "AIAENG","AJANTPHARMA","ALKEM","ANGELONE","APLAPOLLO",
    "APTUS","ASTRAL","ATGL","ATUL","AUROPHARMA",
    "AUBANK","BALKRISIND","BANDHANBNK","BHEL","BIOCON",
    "BLUESTARCO","BRIGADE","CAMS","CANFINHOME","CDSL",
    "CESC","CGPOWER","CHAMBLFERT","COFORGE","CONCOR",
    "CROMPTON","CUMMINSIND","DALBHARAT","DEEPAKNTR","DELHIVERY",
    "DIXON","DLF","ELGIEQUIP","EMAMILTD","ENDURANCE",
    "ESCORTS","EXIDEIND","FEDERALBNK","FIVESTAR","FLUOROCHEM",
    "FORTIS","GLAND","GLAXO","GMRAIRPORT","GODFRYPHLP",
    "GODREJIND","GPIL","GRINDWELL","HAL","HUDCO",
    "IEX","IGL","IIFL","INDHOTEL","INDIGO",
    "IPCALAB","ISEC","JKCEMENT","JSL","JSWENERGY",
    "JUBLFOOD","KAJARIACER","KANSAINER","KARURVYSYA","KAYNES",
    "KEC","KEI","KPITTECH","KRBL","KTKBANK",
    "L&TFH","LALPATHLAB","LATENTVIEW","LICHSGFIN","LINDEINDIA",
    "LUPIN","MANAPPURAM","MAPMYINDIA","MASTEK","MCX",
    "MEDANTA","METROBRAND","MFSL","MGL","MOTILALOFS",
    "MTAR","MUTHOOTFIN","NATCOPHARM","NBCC","NCC",
    "NHPC","NLCINDIA","NSLNISP","OBEROIRLTY","OLECTRA",
    "PATANJALI","PERSISTENT","PHOENIXLTD","PNBHOUSING","POLYCAB",
    "POONAWALLA","PRESTIGE","PVRINOX","RADICO","RAILTEL",
    "RAMCOCEM","RAYMOND","REDINGTON","RELAXO","RITES",
    "RRKABEL","SAFARI","SAREGAMA","SBICARD","SCHAEFFLER",
    "SJVN","SKFINDIA","SOBHA","SPARC","STARHEALTH",
    "STLTECH","SUMICHEM","SUNDARMFIN","SUPREMEIND","SUNTV",
    "SYNGENE","TANLA","TATACOMM","TATAELXSI","TATAINVEST",
    "TCNSBRANDS","TEAMLEASE","TIMKEN","TITAGARH","TORNTPOWER",
    "TRITURBINE","UCOBANK","UJJIVANSFB","UNOMINDA","UTIAMC",
    "VGUARD","VMART","WHIRLPOOL","YESBANK","ZEEL",
    "RKFORGE","ROUTE","INDIANB","ANANDRATHI",
]

def get_hardcoded_fallback():
    combined = NIFTY_50 + NIFTY_NEXT_50 + NIFTY_MIDCAP_150
    seen = set(); result = []
    for s in combined:
        if s not in seen: seen.add(s); result.append(s)
    return result

# ══════════════════════════════════════════════════════════════════════
# CACHE: Save/Load symbols
# ══════════════════════════════════════════════════════════════════════
def save_symbols_cache(symbols):
    os.makedirs(HISTORY_DIR, exist_ok=True)
    with open(SYMBOLS_CACHE_FILE, "w") as f:
        json.dump({
            "symbols":    symbols,
            "count":      len(symbols),
            "saved_at":   datetime.now().strftime("%d %b %Y %H:%M IST"),
            "saved_date": datetime.now().strftime("%Y-%m-%d"),
        }, f, indent=2)
    print(f"  ✓ Cache saved: {len(symbols)} symbols → {SYMBOLS_CACHE_FILE}")

def load_symbols_cache():
    if not os.path.exists(SYMBOLS_CACHE_FILE):
        return None, None
    try:
        with open(SYMBOLS_CACHE_FILE) as f:
            d = json.load(f)
        syms = d.get("symbols", [])
        return (syms, d.get("saved_at")) if len(syms) >= 200 else (None, None)
    except:
        return None, None

# ══════════════════════════════════════════════════════════════════════
# NSE Session
# ══════════════════════════════════════════════════════════════════════
def get_nse_session():
    s = requests.Session()
    s.headers.update(NSE_HEADERS)
    try:
        s.get("https://www.nseindia.com", timeout=15); time.sleep(2)
        s.get("https://www.nseindia.com/market-data/all-reports", timeout=10); time.sleep(1)
    except Exception as e:
        print(f"    Session: {e}")
    return s

# ══════════════════════════════════════════════════════════════════════
# STEP 1: Get Nifty 250 symbols — 3 tier
# ══════════════════════════════════════════════════════════════════════
def fetch_symbols_live(session):
    endpoints = [
        ("LARGEMIDCAP250",
         "https://www.nseindia.com/api/equity-stockIndices?index=NIFTY%20LARGEMIDCAP%20250",
         200),
        ("NIFTY100",
         "https://www.nseindia.com/api/equity-stockIndices?index=NIFTY%20100",
         90),
        ("MIDCAP150",
         "https://www.nseindia.com/api/equity-stockIndices?index=NIFTY%20MIDCAP%20150",
         100),
    ]
    s100 = []; smid = []
    for name, url, minc in endpoints:
        try:
            r = session.get(url, timeout=20)
            if r.status_code != 200:
                print(f"    ⚠ {name}: HTTP {r.status_code}"); time.sleep(1); continue
            syms = [x["symbol"] for x in r.json().get("data",[])
                    if x.get("symbol") and x["symbol"] not in
                    ("NIFTY LARGEMIDCAP 250","NIFTY 100","NIFTY MIDCAP 150","NIFTY50")]
            if len(syms) < minc:
                print(f"    ⚠ {name}: {len(syms)} symbols"); time.sleep(1); continue
            print(f"    ✓ {name}: {len(syms)} symbols")
            if "LARGEMIDCAP" in name and len(syms) >= 200: return syms
            if "100" in name:  s100 = syms
            if "MIDCAP" in name: smid = syms
            time.sleep(1)
        except Exception as e:
            print(f"    ⚠ {name}: {e}")
    if len(s100) >= 90 and len(smid) >= 100:
        combined = list(dict.fromkeys(s100 + smid))
        print(f"    ✓ Combined: {len(combined)} symbols")
        return combined
    return []

def get_nifty250_symbols():
    print("\n[STEP 1] Fetching Nifty 250 symbols (3-tier)…")
    session = get_nse_session()
    sym_source = "Hardcoded (Tier 3)"

    # Tier 1: Live NSE API
    print("  [Tier 1] NSE Live API…")
    try:
        live = fetch_symbols_live(session)
        if len(live) >= 200:
            print(f"  ✅ Tier 1: {len(live)} live symbols")
            save_symbols_cache(live)
            sym_source = "NSE Live API (Tier 1)"
            return live, session, sym_source
        print(f"  ⚠ Tier 1: only {len(live)} symbols")
    except Exception as e:
        print(f"  ⚠ Tier 1 error: {e}")

    # Tier 2: Saved cache
    print("  [Tier 2] Loading saved cache…")
    cached, saved_at = load_symbols_cache()
    if cached:
        print(f"  ✅ Tier 2: {len(cached)} cached symbols (from {saved_at})")
        sym_source = f"Auto Cache (Tier 2) — saved {saved_at}"
        return cached, session, sym_source

    # Tier 3: Hardcoded fallback
    print("  [Tier 3] Using hardcoded fallback…")
    fallback = get_hardcoded_fallback()
    print(f"  ✅ Tier 3: {len(fallback)} hardcoded symbols")
    return fallback, session, sym_source

# ══════════════════════════════════════════════════════════════════════
# STEP 2: Yahoo Finance
# ══════════════════════════════════════════════════════════════════════
def fetch_yfinance(symbols):
    import yfinance as yf

    print(f"\n[STEP 2] Yahoo Finance — {len(symbols)} symbols…")
    end   = datetime.today()
    start = end - timedelta(days=int(HISTORY_DAYS * 1.6))

    tickers    = [to_yf(s) for s in symbols]
    all_close  = {}
    all_volume = {}
    all_today  = {}
    failed     = []
    BATCH      = 50

    for i in range(0, len(tickers), BATCH):
        bt = tickers[i:i+BATCH]
        bs = symbols[i:i+BATCH]
        print(f"  Batch {i//BATCH+1}/{(len(tickers)-1)//BATCH+1}: "
              f"{bs[0]}…{bs[-1]}")
        try:
            raw = yf.download(
                bt,
                start=start.strftime("%Y-%m-%d"),
                end=end.strftime("%Y-%m-%d"),
                auto_adjust=True, progress=False,
                threads=True, group_by="ticker",
            )
            for sym, tk in zip(bs, bt):
                try:
                    if len(bt) == 1:
                        cs=raw["Close"]; vs=raw["Volume"]
                        hs=raw["High"];  ls=raw["Low"]
                    else:
                        if tk not in raw.columns.get_level_values(0):
                            failed.append(sym); continue
                        cs=raw[tk]["Close"]; vs=raw[tk]["Volume"]
                        hs=raw[tk]["High"];  ls=raw[tk]["Low"]
                    cs=cs.dropna(); vs=vs.dropna()
                    if len(cs) < 10:
                        failed.append(sym); continue
                    ld = cs.index[-1]
                    all_close[sym]  = cs
                    all_volume[sym] = vs
                    all_today[sym]  = {
                        "ltp":       round(float(cs.iloc[-1]), 2),
                        "prevClose": round(float(cs.iloc[-2]), 2) if len(cs)>=2 else 0,
                        "high":      round(float(hs.loc[ld]),  2) if ld in hs.index else 0,
                        "low":       round(float(ls.loc[ld]),  2) if ld in ls.index else 0,
                        "volume":    int(float(vs.loc[ld]))        if ld in vs.index else 0,
                        "date":      ld.strftime("%Y-%m-%d"),
                    }
                except Exception as e:
                    failed.append(sym)
        except Exception as e:
            print(f"  ✗ Batch error: {e}")
        time.sleep(1)

    td_str = last_trading_day().strftime("%Y-%m-%d")
    fresh  = sum(1 for v in all_today.values() if v["date"] == td_str)
    total  = len(all_today)

    if failed:
        print(f"  ⚠ Failed ({len(failed)}): {', '.join(failed[:10])}"
              f"{'...' if len(failed)>10 else ''}")
    print(f"  ✓ Fetched: {total} | Today ({td_str}): {fresh} fresh")

    if total == 0 or fresh < total * 0.3:
        print("  ⚠ Insufficient fresh data — switching to Bhavcopy")
        return None, None, None

    return all_close, all_volume, all_today

# ══════════════════════════════════════════════════════════════════════
# STEP 3: NSE Bhavcopy fallback
# ══════════════════════════════════════════════════════════════════════
BHAVCOPY_URL = (
    "https://nsearchives.nseindia.com/content/cm/"
    "BhavCopy_NSE_CM_0_0_0_{date}_F_0000.csv.zip"
)

def download_bhavcopy(date_obj, session):
    ds    = date_obj.strftime("%Y%m%d")
    cache = os.path.join(HISTORY_DIR, f"bhav_{ds}.csv")
    if os.path.exists(cache): return pd.read_csv(cache)
    os.makedirs(HISTORY_DIR, exist_ok=True)
    resp = session.get(BHAVCOPY_URL.format(date=ds), timeout=30)
    if resp.status_code != 200 or len(resp.content) < 500:
        raise Exception(f"HTTP {resp.status_code}")
    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        name = [n for n in z.namelist() if n.endswith(".csv")][0]
        df = pd.read_csv(z.open(name))
    df.to_csv(cache, index=False); return df

def norm(df):
    df.columns = [c.strip().upper().replace(" ","") for c in df.columns]
    return df.rename(columns={
        "TCKRSYMB":"SYMBOL","CLSPRIC":"CLOSE","HGHPRIC":"HIGH",
        "LWPRIC":"LOW","PRVSCLSGPRIC":"PREVCLOSE",
        "TTLTRADGVOL":"VOLUME","TTLTRFVAL":"VALUE","SCTYSRS":"SERIES",
    })

def fetch_bhavcopy(today, symbols, session):
    print(f"\n[STEP 3] NSE Bhavcopy — {len(symbols)} symbols…")
    closes = {}; date = today; count = 0; fails = 0
    while count < 210 and fails < 15:
        try:
            df = norm(download_bhavcopy(date, session))
            sc = "SERIES" if "SERIES" in df.columns else None
            eq = df[df[sc]=="EQ"].copy() if sc else df.copy()
            if "SYMBOL" in eq.columns and "CLOSE" in eq.columns:
                eq["CLOSE"] = pd.to_numeric(eq["CLOSE"], errors="coerce")
                closes[date.strftime("%Y%m%d")] = eq.set_index("SYMBOL")["CLOSE"]
                count += 1; fails = 0
        except Exception as e:
            print(f"    skip {date.strftime('%Y%m%d')}: {e}"); fails += 1
        date -= timedelta(days=1)
        while date.weekday() >= 5: date -= timedelta(days=1)
        time.sleep(0.4)

    price_df = pd.DataFrame(closes).T.sort_index().apply(pd.to_numeric, errors="coerce")
    bhav  = norm(download_bhavcopy(today, session))
    sc    = "SERIES" if "SERIES" in bhav.columns else None
    eq_td = bhav[bhav[sc]=="EQ"].set_index("SYMBOL") if sc else bhav.set_index("SYMBOL")
    for col in ["CLOSE","PREVCLOSE","HIGH","LOW","VOLUME"]:
        if col in eq_td.columns:
            eq_td[col] = pd.to_numeric(eq_td[col], errors="coerce")

    all_close, all_volume, all_today = {}, {}, {}
    for sym in symbols:
        if sym not in eq_td.index: continue
        r = eq_td.loc[sym]
        ltp=float(r.get("CLOSE",0) or 0); prv=float(r.get("PREVCLOSE",0) or 0)
        vol=float(r.get("VOLUME",0) or 0); hgh=float(r.get("HIGH",0) or 0)
        low=float(r.get("LOW",0)   or 0)
        if sym in price_df.columns:
            all_close[sym]  = price_df[sym].dropna()
            all_volume[sym] = pd.Series([vol], index=[today])
        all_today[sym] = {"ltp":round(ltp,2),"prevClose":round(prv,2),
                          "high":round(hgh,2),"low":round(low,2),
                          "volume":int(vol),"date":today.strftime("%Y-%m-%d")}
    print(f"  ✓ Bhavcopy: {len(all_today)} stocks")
    return all_close, all_volume, all_today

# ══════════════════════════════════════════════════════════════════════
# STEP 4: Screen — DMA + signals
# ══════════════════════════════════════════════════════════════════════
def screen(all_close, all_volume, all_today, symbols):
    print(f"\n[STEP 4] Screening {len(all_today)} stocks…")
    records = []
    for sym in symbols:
        if sym not in all_today: continue
        td=all_today[sym]; ltp=td["ltp"]; prv=td["prevClose"]
        vol=td["volume"]; hgh=td.get("high",0); low=td.get("low",0)
        chg=round(ltp-prv,2); chg_pct=round((chg/prv*100) if prv else 0,2)
        hist=all_close.get(sym,pd.Series(dtype=float)).dropna()
        hd=hist.iloc[:-1] if len(hist)>1 else hist
        dma50  = round(float(hd.tail(50).mean()),  2) if len(hd)>=50  else None
        dma100 = round(float(hd.tail(100).mean()), 2) if len(hd)>=100 else None
        dma200 = round(float(hd.tail(200).mean()), 2) if len(hd)>=200 else None
        vh=all_volume.get(sym,pd.Series(dtype=float)).dropna()
        avg_vol=float(vh.iloc[:-1].tail(20).mean()) if len(vh)>=5 else 0
        val_cr=round(ltp*vol/1e7,2)
        a50  = bool(ltp>dma50)  if dma50  else False
        a100 = bool(ltp>dma100) if dma100 else False
        a200 = bool(ltp>dma200) if dma200 else False
        pct200=round((ltp-dma200)/dma200*100,2) if dma200 else None
        hi_vol=bool(vol>avg_vol*HIGH_VOL_MULTIPLIER) if avg_vol else False
        hi_val=bool(val_cr>=HIGH_VALUE_CRORE)
        over  =bool(pct200>DMA200_OVEREXTENDED_PCT) if pct200 is not None else False
        if   a50 and a100 and a200 and hi_vol and not over: sig="BUY"
        elif a50 and a100 and a200 and over:                sig="OVEREXTENDED"
        elif a50 and a100 and a200:                         sig="HOLD"
        else:                                               sig="WATCH"
        records.append({
            "symbol":sym,"ltp":ltp,"prevClose":prv,"change":chg,"changePct":chg_pct,
            "high":hgh,"low":low,"volume":vol,"avgVolume":int(avg_vol),
            "highVolume":hi_vol,"valueCr":val_cr,"highValue":hi_val,
            "dma50":dma50,"aboveDMA50":a50,
            "dma100":dma100,"aboveDMA100":a100,
            "dma200":dma200,"aboveDMA200":a200,
            "pctAboveDMA200":pct200,
            "dma200Alert":"OVEREXTENDED" if over else "IN_RANGE",
            "signal":sig,"dataDate":td["date"],
        })
    df=pd.DataFrame(records)
    df=df.sort_values("pctAboveDMA200",ascending=False,
                      key=lambda x:pd.to_numeric(x,errors="coerce"))
    print(f"\n  ✅ Screened: {len(df)}/{len(symbols)}")
    print(f"     BUY:{(df.signal=='BUY').sum()}  "
          f"HOLD:{(df.signal=='HOLD').sum()}  "
          f"OVEREXTENDED:{(df.signal=='OVEREXTENDED').sum()}  "
          f"WATCH:{(df.signal=='WATCH').sum()}")
    return df

# ══════════════════════════════════════════════════════════════════════
# STEP 5: Export
# ══════════════════════════════════════════════════════════════════════
def export_json(df, method, total_syms, sym_source):
    os.makedirs(DOCS_DIR, exist_ok=True)
    payload={
        "generated_at":  datetime.now().strftime("%d %b %Y %H:%M IST"),
        "trading_day":   df["dataDate"].iloc[0] if not df.empty else "N/A",
        "data_source":   method,
        "symbol_source": sym_source,
        "index_total":   total_syms,
        "fetched_total": len(df),
        "stocks": df.where(pd.notnull(df),None).to_dict(orient="records"),
    }
    path=os.path.join(DOCS_DIR,"screener_data.json")
    with open(path,"w") as f: json.dump(payload,f,separators=(",",":"))
    print(f"  ✓ JSON  → {path} ({os.path.getsize(path)/1024:.1f} KB)")

CLR=dict(header="0d1e3b",buy="C6EFCE",hold="FFEB9C",
         over="FFC7CE",watch="F2F2F2",yes="E2EFDA",no="FCE4D6")

def export_excel(df):
    rows=[]
    for _,s in df.iterrows():
        rows.append({
            "Symbol":s.symbol,"Date":s.dataDate,
            "LTP (₹)":s.ltp,"Prev Close (₹)":s.prevClose,
            "Change (₹)":s.change,"Change (%)":s.changePct,
            "High (₹)":s.high,"Low (₹)":s.low,
            "Volume":s.volume,"Avg Vol (20D)":s.avgVolume,
            "High Volume":"YES" if s.highVolume else "NO",
            "Value (₹ Cr)":s.valueCr,
            "High Value":"YES" if s.highValue else "NO",
            "DMA 50 (₹)":s.dma50,"Above DMA 50":"YES" if s.aboveDMA50 else "NO",
            "DMA 100 (₹)":s.dma100,"Above DMA 100":"YES" if s.aboveDMA100 else "NO",
            "DMA 200 (₹)":s.dma200,"Above DMA 200":"YES" if s.aboveDMA200 else "NO",
            "% vs DMA 200":s.pctAboveDMA200,
            "DMA 200 Alert":"⚠ OVEREXTENDED (>10%)" if s.dma200Alert=="OVEREXTENDED"
                            else "✓ Within Range",
            "Signal":s.signal,
        })
    out=pd.DataFrame(rows)
    os.makedirs(DOCS_DIR,exist_ok=True)
    arc=os.path.join(DOCS_DIR,"archive"); os.makedirs(arc,exist_ok=True)
    ds=datetime.today().strftime("%Y-%m-%d")
    dated=os.path.join(arc,f"NSE_DMA_Screener_{ds}.xlsx")
    latest=os.path.join(DOCS_DIR,"latest.xlsx")
    def write(path):
        with pd.ExcelWriter(path,engine="openpyxl") as w:
            out.to_excel(w,sheet_name="All Nifty 250",index=False)
            out[out.Signal.isin(["BUY","HOLD","OVEREXTENDED"])].to_excel(w,sheet_name="Screened",index=False)
            out[out.Signal=="BUY"].to_excel(w,sheet_name="BUY Signals",index=False)
        wb=load_workbook(path)
        for ws in wb.worksheets: _fmt(ws)
        wb.save(path)
    write(dated)
    import shutil; shutil.copy(dated,latest)
    print(f"  ✓ Excel → {latest}")

def _fmt(ws):
    thin=Side(style="thin",color="D0D0D0")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    for c in ws[1]:
        c.fill=PatternFill("solid",fgColor=CLR["header"])
        c.font=Font(bold=True,color="FFFFFF",size=10)
        c.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.row_dimensions[1].height=30
    for col in ws.columns:
        mw=max((len(str(c.value or "")) for c in col),default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(mw+4,24)
    hdrs={c.value:c.column for c in ws[1]}
    SC={"BUY":CLR["buy"],"HOLD":CLR["hold"],"OVEREXTENDED":CLR["over"],"WATCH":CLR["watch"]}
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.border=bdr; cell.alignment=Alignment(horizontal="right")
        for cn in ["Above DMA 50","Above DMA 100","Above DMA 200"]:
            ci=hdrs.get(cn)
            if ci:
                c=row[ci-1]
                c.fill=PatternFill("solid",fgColor=(CLR["yes"] if c.value=="YES" else CLR["no"]))
                c.font=Font(bold=True)
        ci=hdrs.get("Signal")
        if ci:
            c=row[ci-1]; cl=SC.get(str(c.value),"")
            if cl: c.fill=PatternFill("solid",fgColor=cl); c.font=Font(bold=True)
        ci=hdrs.get("DMA 200 Alert")
        if ci:
            c=row[ci-1]
            c.fill=PatternFill("solid",fgColor=(CLR["over"] if "OVER" in str(c.value) else CLR["buy"]))
            c.font=Font(bold=True)
        ci=hdrs.get("Change (%)")
        if ci:
            c=row[ci-1]
            try: v=float(c.value); c.font=Font(color=("375623" if v>=0 else "9C0006"),bold=True)
            except: pass
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions

def last_trading_day():
    d=datetime.today()
    while d.weekday()>=5: d-=timedelta(days=1)
    return d

# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
def run():
    print("="*60)
    print("  NSE Nifty LargeMidCap 250 — Daily DMA Screener")
    print(f"  {datetime.now().strftime('%d %b %Y %H:%M UTC')}")
    print("="*60)

    today = last_trading_day()
    print(f"\n  Trading day: {today.strftime('%d %b %Y (%A)')}")

    # Step 1: Symbols
    symbols, session, sym_source = get_nifty250_symbols()
    print(f"\n  Symbol source : {sym_source}")
    print(f"  Symbol count  : {len(symbols)}")

    # Step 2: Yahoo Finance
    data_method="Yahoo Finance"
    all_close,all_volume,all_today=None,None,None
    try:
        all_close,all_volume,all_today=fetch_yfinance(symbols)
    except Exception as e:
        print(f"  ✗ yfinance: {e}")

    # Step 3: Bhavcopy fallback
    if not all_today:
        data_method="NSE Bhavcopy"
        print("\n  → Switching to NSE Bhavcopy…")
        try:
            all_close,all_volume,all_today=fetch_bhavcopy(today,symbols,session)
        except Exception as e:
            print(f"  ✗ Bhavcopy: {e}"); raise SystemExit(1)

    if not all_today:
        print("  ✗ FATAL: No data."); raise SystemExit(1)

    # Step 4: Screen
    result_df=screen(all_close,all_volume,all_today,symbols)
    if result_df.empty:
        print("  ✗ FATAL: No results."); raise SystemExit(1)

    # Step 5: Export
    print("\n[STEP 5] Exporting…")
    export_json(result_df,data_method,len(symbols),sym_source)
    export_excel(result_df)

    print("\n"+"="*60)
    print(f"  ✅ COMPLETE!")
    print(f"  Symbol source : {sym_source}")
    print(f"  Data source   : {data_method}")
    print(f"  Index symbols : {len(symbols)}")
    print(f"  Stocks fetched: {len(result_df)}")
    print("="*60)

if __name__=="__main__":
    run()

